"""
Gerador de Output — LEX FINDER
Produz dois artefatos a partir dos lançamentos filtrados:
  1. Planilha Excel (.xlsx) com os descontos identificados
  2. PDF consolidado (todos os extratos em ordem cronológica)

Dependências: openpyxl, pypdf
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfWriter, PdfReader
    PYPDF_OK = True
except ImportError:
    try:
        from PyPDF2 import PdfWriter, PdfReader
        PYPDF_OK = True
    except ImportError:
        PYPDF_OK = False


# ── Paleta de cores ───────────────────────────────────────────────────────────

COR_CABECALHO = "1E3A5F"      # Azul escuro
COR_CABECALHO_FONTE = "FFFFFF"
COR_LINHA_PAR = "EBF3FB"      # Azul claro alternado
COR_TOTAL = "D4E6F1"
COR_DESTAQUE = "FFC300"       # Amarelo para valores altos


# ── Formatadores ──────────────────────────────────────────────────────────────

def _parse_data(data_str: str) -> datetime | None:
    """Converte 'DD/MM/AAAA' para datetime."""
    try:
        return datetime.strptime(data_str.strip(), "%d/%m/%Y")
    except (ValueError, AttributeError):
        return None


def _brl(valor: float | None) -> str:
    """Formata float como BRL: 1234.56 → 'R$ 1.234,56'"""
    if valor is None:
        return "—"
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ── Gerador Excel ─────────────────────────────────────────────────────────────

def gerar_excel(
    lancamentos_filtrados: list[dict],
    caminho_saida: str,
    meta_banco: dict | None = None,
) -> str:
    """
    Gera planilha Excel com os lançamentos filtrados (descontos identificados).

    lancamentos_filtrados: lista de dicts com campos:
        data, historico, docto, debito, credito, saldo, pagina, banco,
        titulo_desconto, rubrica_matched

    Retorna o caminho do arquivo gerado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descontos Identificados"

    # ── Estilos reutilizáveis ──
    fonte_cabecalho = Font(bold=True, color=COR_CABECALHO_FONTE, size=11)
    fill_cabecalho = PatternFill("solid", fgColor=COR_CABECALHO)
    fill_par = PatternFill("solid", fgColor=COR_LINHA_PAR)
    fill_total = PatternFill("solid", fgColor=COR_TOTAL)
    alinha_centro = Alignment(horizontal="center", vertical="center")
    alinha_dir = Alignment(horizontal="right", vertical="center")
    alinha_esq = Alignment(horizontal="left", vertical="center")

    borda_fina = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Cabeçalho do relatório ──
    linha = 1
    if meta_banco:
        ws.merge_cells(f"A{linha}:H{linha}")
        cel = ws.cell(linha, 1, "LEX FINDER — Relatório de Descontos Indevidos")
        cel.font = Font(bold=True, size=14, color=COR_CABECALHO)
        cel.alignment = alinha_centro
        linha += 1

        infos = []
        if meta_banco.get("cliente"):
            infos.append(f"Cliente: {meta_banco['cliente']}")
        if meta_banco.get("banco"):
            infos.append(f"Banco: {meta_banco['banco']}")
        if meta_banco.get("periodo"):
            infos.append(f"Período: {meta_banco['periodo']}")
        infos.append(f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        ws.merge_cells(f"A{linha}:H{linha}")
        cel = ws.cell(linha, 1, "  |  ".join(infos))
        cel.font = Font(italic=True, color="666666", size=10)
        cel.alignment = alinha_centro
        linha += 1

    linha += 1  # linha em branco

    # ── Linha de cabeçalho da tabela ──
    cabecalhos = [
        "Data", "Categoria", "Histórico", "Doc.",
        "Débito (R$)", "Crédito (R$)", "Rubrica Matched", "Pág."
    ]
    larguras = [12, 25, 45, 12, 14, 14, 30, 6]

    for col, (texto, larg) in enumerate(zip(cabecalhos, larguras), 1):
        cel = ws.cell(linha, col, texto)
        cel.font = fonte_cabecalho
        cel.fill = fill_cabecalho
        cel.alignment = alinha_centro
        cel.border = borda_fina
        ws.column_dimensions[get_column_letter(col)].width = larg

    linha_cabecalho = linha
    linha += 1

    # ── Dados ──
    total_debito = 0.0
    total_credito = 0.0

    # Ordenar por data
    def _sort_key(l):
        d = _parse_data(l.get("data", ""))
        return d if d else datetime.min

    dados = sorted(lancamentos_filtrados, key=_sort_key)

    for idx, l in enumerate(dados):
        fill = fill_par if idx % 2 == 0 else PatternFill()

        valores_linha = [
            l.get("data", ""),
            l.get("titulo_desconto", ""),
            l.get("historico", ""),
            l.get("docto", "") or "—",
            l.get("debito") or None,
            l.get("credito") or None,
            l.get("rubrica_matched", ""),
            l.get("pagina", ""),
        ]

        for col, valor in enumerate(valores_linha, 1):
            cel = ws.cell(linha, col, valor)
            cel.fill = fill
            cel.border = borda_fina

            if col in {5, 6} and isinstance(valor, float):
                cel.number_format = '#,##0.00'
                cel.alignment = alinha_dir
                if col == 5 and valor:
                    total_debito += valor
                if col == 6 and valor:
                    total_credito += valor
            elif col in {1, 8}:
                cel.alignment = alinha_centro
            else:
                cel.alignment = alinha_esq

        linha += 1

    # ── Linha de total ──
    linha += 1
    ws.cell(linha, 4, "TOTAL").font = Font(bold=True)
    ws.cell(linha, 4).fill = fill_total
    ws.cell(linha, 4).alignment = alinha_dir

    cel_tot_deb = ws.cell(linha, 5, total_debito)
    cel_tot_deb.font = Font(bold=True, color="C0392B")
    cel_tot_deb.fill = fill_total
    cel_tot_deb.number_format = '#,##0.00'
    cel_tot_deb.alignment = alinha_dir

    cel_tot_cred = ws.cell(linha, 6, total_credito)
    cel_tot_cred.font = Font(bold=True, color="1A7A4A")
    cel_tot_cred.fill = fill_total
    cel_tot_cred.number_format = '#,##0.00'
    cel_tot_cred.alignment = alinha_dir

    # ── Aba de resumo por categoria ──
    ws_resumo = wb.create_sheet("Resumo por Categoria")
    _gerar_aba_resumo(ws_resumo, dados, fonte_cabecalho, fill_cabecalho, borda_fina)

    # ── Salvar ──
    caminho = Path(caminho_saida)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)

    return str(caminho)


def _gerar_aba_resumo(ws, dados, fonte_cabecalho, fill_cabecalho, borda_fina):
    """Aba com totais agrupados por categoria de desconto."""
    from collections import defaultdict

    totais = defaultdict(float)
    contagens = defaultdict(int)

    for l in dados:
        cat = l.get("titulo_desconto", "Desconhecido")
        debito = l.get("debito") or 0.0
        totais[cat] += debito
        contagens[cat] += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16

    linha = 1
    for col, texto in enumerate(["Categoria", "Qtd.", "Total Débito (R$)"], 1):
        cel = ws.cell(linha, col, texto)
        cel.font = fonte_cabecalho
        cel.fill = fill_cabecalho
        cel.border = borda_fina
        cel.alignment = Alignment(horizontal="center")
    linha += 1

    fill_par = PatternFill("solid", fgColor="EBF3FB")

    for idx, (cat, total) in enumerate(sorted(totais.items(), key=lambda x: -x[1])):
        fill = fill_par if idx % 2 == 0 else PatternFill()
        ws.cell(linha, 1, cat).fill = fill
        ws.cell(linha, 1).border = borda_fina

        cel_qtd = ws.cell(linha, 2, contagens[cat])
        cel_qtd.fill = fill
        cel_qtd.border = borda_fina
        cel_qtd.alignment = Alignment(horizontal="center")

        cel_val = ws.cell(linha, 3, total)
        cel_val.fill = fill
        cel_val.border = borda_fina
        cel_val.number_format = '#,##0.00'
        cel_val.alignment = Alignment(horizontal="right")

        linha += 1


# ── Gerador PDF consolidado ───────────────────────────────────────────────────

def mesclar_pdfs(
    lista_pdfs: list[str],
    caminho_saida: str,
    ordenar_por_nome: bool = True,
) -> str:
    """
    Mescla múltiplos PDFs em um único arquivo na ordem informada.
    Se ordenar_por_nome=True, ordena os arquivos alfabeticamente (útil quando
    o nome inclui data: 2023-01.pdf, 2023-02.pdf...).

    Retorna o caminho do arquivo gerado.
    """
    if not PYPDF_OK:
        raise ImportError(
            "pypdf não instalado. Execute: pip install pypdf"
        )

    if ordenar_por_nome:
        lista_pdfs = sorted(lista_pdfs)

    writer = PdfWriter()

    for pdf_path in lista_pdfs:
        try:
            reader = PdfReader(str(pdf_path))
            for page in reader.pages:
                writer.add_page(page)
        except Exception as e:
            print(f"  [AVISO] Não foi possível incluir {pdf_path}: {e}")

    caminho = Path(caminho_saida)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    with open(caminho, "wb") as f:
        writer.write(f)

    return str(caminho)


# ── Teste rápido ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Dados de teste simulados
    lancamentos_teste = [
        {
            "data": "05/03/2024",
            "historico": "TARIFA BANCARIA CESTA B.EXPRESSO4",
            "docto": "000001",
            "debito": 49.90,
            "credito": None,
            "saldo": 1000.00,
            "pagina": 1,
            "banco": "Bradesco",
            "titulo_desconto": "Tarifas Bancárias",
            "rubrica_matched": "TARIFA BANCARIA",
        },
        {
            "data": "10/03/2024",
            "historico": "MORA CREDITO PESSOAL 9990026",
            "docto": "000002",
            "debito": 125.40,
            "credito": None,
            "saldo": 874.60,
            "pagina": 1,
            "banco": "Bradesco",
            "titulo_desconto": "Mora e Juros de Empréstimo",
            "rubrica_matched": "MORA CREDITO",
        },
        {
            "data": "15/03/2024",
            "historico": "ENCARGOS LIMITE DE CRED ENCARGO - 07,73%",
            "docto": "000003",
            "debito": 87.20,
            "credito": None,
            "saldo": 787.40,
            "pagina": 2,
            "banco": "Bradesco",
            "titulo_desconto": "Encargos de Limite de Crédito",
            "rubrica_matched": "ENCARGOS LIMITE DE CRED",
        },
    ]

    meta = {
        "cliente": "NICOLAS GOMES ADVOCACIA",
        "banco": "Bradesco",
        "periodo": "01/01/2024 a 31/03/2024",
    }

    saida = gerar_excel(lancamentos_teste, "output/teste_relatorio.xlsx", meta)
    print(f"Excel gerado: {saida}")
    print(f"Total registros: {len(lancamentos_teste)}")
